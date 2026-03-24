import streamlit as st
import numpy as np
import pandas as pd
import math
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from fpdf import FPDF

# Constants
TEMPERATURE_COEFFICIENT = 0.000025
COVERAGE_FACTOR = 2
BMC_FLOOR = 0.083  # Best Measurement Capability floor


def calculate_v_eff(U_c, U1, V1):
    """Calculate Effective Degrees of Freedom using the Welch-Satterthwaite equation.

    The general formula is:
        V_eff = U_c^4 / sum(U_i^4 / V_i)

    Because V2 through V6 are infinite, each corresponding term (U_i^4 / V_i)
    mathematically evaluates to zero and contributes nothing to the denominator.
    The sum therefore reduces to just the U1 term, giving:
        V_eff = U_c^4 / (U1^4 / V1)
    """
    denominator = (U1**4) / V1
    if denominator == 0:
        return math.inf
    return (U_c**4) / denominator

st.set_page_config(page_title="MTSL Uncertainty Calc", layout="wide")

st.title("⚡ Uncertainty Calculation Worksheet - MTSL Palakkad")
st.caption("Meter Testing & Standards Laboratory, Palakkad")
st.markdown("---")

# Sidebar for Reference/Certificate Settings
st.sidebar.header("📋 Configuration Parameters")

st.sidebar.markdown("### Calibration Details")

calibration_date = st.sidebar.date_input(
    "Date of Calibration",
    value=None,
    help="Enter the date of calibration"
)

temp_cal_certificate = st.sidebar.number_input(
    "Temperature as per Calibration Certificate of Reference Standard (°C)",
    value=25.0,
    format="%.2f",
    help="Temperature recorded in the calibration certificate of the reference standard"
)

room_temp_calibration = st.sidebar.number_input(
    "Room Temperature During Calibration (°C)",
    value=25.0,
    format="%.2f",
    help="Actual room temperature during the calibration process"
)

st.sidebar.markdown("### Reference Standard Details")

ref_make = st.sidebar.text_input(
    "Make",
    value="",
    key="ref_make",
    help="Manufacturer/make of the reference standard"
)

ref_model = st.sidebar.text_input(
    "Model",
    value="",
    key="ref_model",
    help="Model of the reference standard"
)

ref_sl_no = st.sidebar.text_input(
    "Sl No.",
    value="",
    key="ref_sl_no",
    help="Serial number of the reference standard"
)

ref_accuracy_class = st.sidebar.text_input(
    "Accuracy Class",
    value="",
    key="ref_accuracy_class",
    help="Accuracy class of the reference standard"
)

ref_due_date = st.sidebar.date_input(
    "Date of Due Calibration",
    value=None,
    key="ref_due_date",
    help="Due date of calibration for the reference standard"
)

st.sidebar.markdown("### Device Under Calibration Details")

duc_make = st.sidebar.text_input(
    "Make",
    value="",
    key="duc_make",
    help="Manufacturer/make of the device under calibration"
)

duc_model = st.sidebar.text_input(
    "Model",
    value="",
    key="duc_model",
    help="Model of the device under calibration"
)

duc_sl_no = st.sidebar.text_input(
    "Sl No.",
    value="",
    key="duc_sl_no",
    help="Serial number of the device under calibration"
)

duc_accuracy_class = st.sidebar.text_input(
    "Accuracy Class",
    value="",
    key="duc_accuracy_class",
    help="Accuracy class of the device under calibration"
)

duc_due_date = st.sidebar.date_input(
    "Date of Due Calibration",
    value=None,
    key="duc_due_date",
    help="Due date of calibration for the device under calibration"
)

st.sidebar.markdown("### Reference & Certificate")

ref_standard_accuracy = st.sidebar.number_input(
    "Reference Standard Accuracy (%)",
    value=0.05,
    format="%.4f",
    help="Enter the reference standard accuracy percentage"
)

certificate_uncertainty = st.sidebar.number_input(
    "Certificate Uncertainty (%)",
    value=0.03,
    format="%.4f",
    help="Enter the certificate uncertainty percentage"
)

duc_resolution = st.sidebar.number_input(
    "DUC Resolution",
    value=0.001,
    format="%.4f",
    help="Enter the DUC (Device Under Calibration) resolution"
)

st.sidebar.markdown("### Environmental & Drift Parameters")

temp_difference = round(temp_cal_certificate - room_temp_calibration, 2)
st.sidebar.metric(
    "Temperature Difference (°C)",
    f"{temp_difference:.2f}",
    help="Automatically calculated as: Temperature as per Calibration Certificate of Reference Standard − Room Temperature During Calibration"
)

age_factor = st.sidebar.number_input(
    "Age Factor (%/year)",
    value=0.00001,
    format="%.7f",
    step=0.000001,
    help="Enter the drift percentage per year for the reference standard"
)

years_in_service = st.sidebar.number_input(
    "Years in Service",
    value=10.0,
    step=1.0,
    format="%.1f",
    help="Enter the total years the reference standard has been in service"
)

# Main area for error readings
st.header("📊 Error Readings Input")
st.markdown("Enter the 10 error readings from your measurements:")

# Initialize session state for readings if not exists
if 'readings_df' not in st.session_state:
    st.session_state.readings_df = pd.DataFrame({
        'Reading #': [f"Reading {i}" for i in range(1, 11)],
        'Error Value': [0.0] * 10
    })

# Use data_editor for easy data entry
edited_df = st.data_editor(
    st.session_state.readings_df,
    use_container_width=True,
    hide_index=True,
    num_rows="fixed",
    column_config={
        "Reading #": st.column_config.TextColumn(
            "Reading #",
            disabled=True,
            width="medium"
        ),
        "Error Value": st.column_config.NumberColumn(
            "Error Value",
            format="%.4f",
            width="medium"
        )
    },
    key="error_readings_editor"
)

# Extract error readings from the edited dataframe
error_readings = edited_df['Error Value'].tolist()

st.markdown("---")

# Electrical Parameters Section
st.header("⚡ Electrical Parameters")
st.markdown("Enter the electrical parameters for power calculation:")

col_elec1, col_elec2 = st.columns(2)

with col_elec1:
    ac_type = st.radio(
        "AC Type",
        options=["Single-phase (1φ)", "Three-phase (3φ)"],
        index=0,
        horizontal=True,
        help="Select the AC system type"
    )
    
    measurement_mode = st.selectbox(
        "Measurement Mode",
        options=["Active", "Reactive", "Apparent"],
        index=0,
        help="Select the measurement mode: Active (kW/kWh), Reactive (kVAr/kVArh), or Apparent (kVA/kVAh)"
    )
    
    voltage = st.number_input(
        "Voltage (V)",
        min_value=0.0,
        value=230.0,
        format="%.2f",
        help="Enter the voltage in volts (V). For three-phase, use line-to-line voltage (V_LL)"
    )
    
with col_elec2:
    current = st.number_input(
        "Current (A)",
        min_value=0.0,
        value=5.0,
        format="%.3f",
        help="Enter the current in amperes (A)"
    )
    
    pf_type = st.selectbox(
        "Power Factor Type",
        options=["Lag", "Lead", "Unity"],
        index=0,
        help="Select whether the power factor is lagging, leading, or unity"
    )
    
    power_factor = st.number_input(
        "Power Factor",
        min_value=0.0,
        max_value=1.0,
        value=0.85,
        format="%.3f",
        help="Enter the power factor magnitude (0.0 to 1.0)"
    )

time_hours = st.number_input(
    "Time Duration (hours)",
    min_value=0.0,
    value=1.0,
    format="%.2f",
    help="Enter the time duration in hours for energy calculation"
)

# Calculate Real Power
if ac_type == "Single-phase (1φ)":
    real_power_w = voltage * current * power_factor
    power_formula = "P(W) = V × I × PF"
else:  # Three-phase
    real_power_w = math.sqrt(3) * voltage * current * power_factor
    power_formula = "P(W) = √3 × V_LL × I × PF"

real_power_kw = real_power_w / 1000.0

# Calculate sin(φ) from power factor magnitude
# Leading (capacitive) load: reactive power is negative; Unity: reactive power is zero
if pf_type == "Unity":
    sin_phi = 0.0
else:
    sin_phi = math.sqrt(max(0.0, 1 - power_factor**2))
    if pf_type == "Lead":
        sin_phi = -sin_phi

# Calculate Reactive Power
if ac_type == "Single-phase (1φ)":
    reactive_power_var = voltage * current * sin_phi
    reactive_formula = "Q(VAr) = V × I × sin(φ)"
else:  # Three-phase
    reactive_power_var = math.sqrt(3) * voltage * current * sin_phi
    reactive_formula = "Q(VAr) = √3 × V_LL × I × sin(φ)"

reactive_power_kvar = reactive_power_var / 1000.0

# Calculate Apparent Power
if ac_type == "Single-phase (1φ)":
    apparent_power_va = voltage * current
    apparent_formula = "S(VA) = V × I"
else:  # Three-phase
    apparent_power_va = math.sqrt(3) * voltage * current
    apparent_formula = "S(VA) = √3 × V_LL × I"

apparent_power_kva = apparent_power_va / 1000.0

# Calculate Energy
real_energy_kwh = real_power_kw * time_hours
reactive_energy_kvarh = reactive_power_kvar * time_hours
apparent_energy_kvah = apparent_power_kva * time_hours

# Display Power & Energy Calculations
st.subheader("⚡ Power & Energy Calculations")
st.info(f"**Measurement Mode: {measurement_mode}**")

col_power1, col_power2 = st.columns(2)

# Determine label suffix based on selected measurement mode
_active_marker = " ✅ (Selected)" if measurement_mode == "Active" else ""
_reactive_marker = " ✅ (Selected)" if measurement_mode == "Reactive" else ""
_apparent_marker = " ✅ (Selected)" if measurement_mode == "Apparent" else ""

with col_power1:
    st.metric(
        f"Real Power (kW){_active_marker}",
        f"{real_power_kw:.3f}",
        help=f"Calculated using: {power_formula}"
    )
    
    st.metric(
        f"Reactive Power (kVAr){_reactive_marker}",
        f"{reactive_power_kvar:.3f}",
        help=f"Calculated using: {reactive_formula}"
    )
    
    st.metric(
        f"Apparent Power (kVA){_apparent_marker}",
        f"{apparent_power_kva:.3f}",
        help=f"Calculated using: {apparent_formula}"
    )

with col_power2:
    st.metric(
        f"Real Energy (kWh){_active_marker}",
        f"{real_energy_kwh:.3f}",
        help=f"Energy = Power × Time = {real_power_kw:.3f} kW × {time_hours:.2f} h"
    )
    
    st.metric(
        f"Reactive Energy (kVArh){_reactive_marker}",
        f"{reactive_energy_kvarh:.3f}",
        help=f"Reactive Energy = Reactive Power × Time = {reactive_power_kvar:.3f} kVAr × {time_hours:.2f} h"
    )
    
    st.metric(
        f"Apparent Energy (kVAh){_apparent_marker}",
        f"{apparent_energy_kvah:.3f}",
        help=f"Apparent Energy = Apparent Power × Time = {apparent_power_kva:.3f} kVA × {time_hours:.2f} h"
    )

st.markdown("---")

# Calculations
st.header("🔬 Uncertainty Analysis Results")

# Calculate uncertainty components
error_array = np.array(error_readings)

# U1 - Repeatability (Standard Deviation / sqrt(n))
U1 = np.std(error_array, ddof=1) / math.sqrt(len(error_array))

# U2 - Reference Standard
U2 = ref_standard_accuracy / math.sqrt(3)

# U3 - Certificate
U3 = certificate_uncertainty / 2

# U4 - Resolution
U4 = duc_resolution / (2 * math.sqrt(3))

# U5 - Temperature Drift
U5 = (TEMPERATURE_COEFFICIENT * temp_difference) / math.sqrt(3)

# U6 - Energy Drift (age-based)
total_drift = age_factor * years_in_service
U6 = total_drift / math.sqrt(3)

# Degrees of Freedom for each uncertainty component
# V1: Type A repeatability based on 10 readings → n-1 = 9
V1 = 9
# V2–V6: Type B components modelled by well-known distributions →
# infinite degrees of freedom by convention
V2 = math.inf
V3 = math.inf
V4 = math.inf
V5 = math.inf
V6 = math.inf


# Average Error
average_error = np.mean(error_array)

# Combined Uncertainty (uc)
uc = math.sqrt(U1**2 + U2**2 + U3**2 + U4**2 + U5**2 + U6**2)

# Effective Degrees of Freedom (Welch-Satterthwaite)
V_eff = calculate_v_eff(uc, U1, V1)

# Expanded Uncertainty (U)
expanded_uncertainty = uc * COVERAGE_FACTOR

# Apply BMC floor
final_expanded_uncertainty = max(expanded_uncertainty, BMC_FLOOR)
bmc_applied = expanded_uncertainty < BMC_FLOOR

# Create Uncertainty Budget Table
st.subheader("📋 Uncertainty Budget")

uncertainty_budget = pd.DataFrame({
    "Component": [
        "U1",
        "U2",
        "U3",
        "U4",
        "U5",
        "U6"
    ],
    "Description": [
        "Repeatability",
        "Reference Standard Accuracy (uncertainty caused due to the accuracy factor of reference standard used for calibration)",
        "Certificate Uncertainty (uncertainty caused due to the uncertainty reported by the lab where reference standard was calibrated)",
        "Resolution",
        "Temperature Drift (uncertainty caused due to the fact that temperature maintained during the calibration by the lab where reference standard was calibrated is different from the reference temperature)",
        "Energy Drift (uncertainty included to account for age of the reference standard)"
    ],
    "Type": ["A", "B", "B", "B", "B", "B"],
    "Distribution": [
        "Normal",
        "Rectangular",
        "Normal (k=2)",
        "Rectangular",
        "Rectangular",
        "Rectangular"
    ],
    "Value": [
        np.std(error_array, ddof=1),
        ref_standard_accuracy,
        certificate_uncertainty,
        duc_resolution,
        TEMPERATURE_COEFFICIENT * temp_difference,
        total_drift
    ],
    "Divisor": ["√10", "√3", "2", "2√3", "√3", "√3"],
    "Standard Uncertainty (ui)": [U1, U2, U3, U4, U5, U6],
    "Degrees of Freedom (ν)": [str(V1), "∞", "∞", "∞", "∞", "∞"]
})

st.dataframe(
    uncertainty_budget,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Component": st.column_config.TextColumn("Component", width="medium"),
        "Description": st.column_config.TextColumn("Description", width="large"),
        "Type": st.column_config.TextColumn("Type", width="small"),
        "Distribution": st.column_config.TextColumn("Distribution", width="medium"),
        "Value": st.column_config.NumberColumn("Value", format="%.6f", width="medium"),
        "Divisor": st.column_config.TextColumn("Divisor", width="small"),
        "Standard Uncertainty (ui)": st.column_config.NumberColumn("Standard Uncertainty (ui)", format="%.6f", width="medium"),
        "Degrees of Freedom (ν)": st.column_config.TextColumn("Degrees of Freedom (ν)", width="medium")
    }
)

st.markdown("---")

# Display results
col_results1, col_results2 = st.columns(2)

with col_results1:
    st.subheader("📈 Uncertainty Components")
    
    st.metric("U1 - Repeatability (Std Dev/√n)", f"{U1:.6f}")
    st.metric("U2 - Reference Standard", f"{U2:.6f}")
    st.metric("U3 - Certificate", f"{U3:.6f}")
    st.metric("U4 - Resolution", f"{U4:.6f}")
    st.metric("U5 - Temperature Drift", f"{U5:.6f}")
    st.metric("U6 - Energy Drift (Age)", f"{U6:.6f}")

with col_results2:
    st.subheader("📊 Final Results")
    
    st.metric(
        "Average Error", 
        f"{average_error:.6f}",
        help="Mean of the 10 error readings"
    )
    
    st.metric(
        "Combined Uncertainty (uc)", 
        f"{uc:.6f}",
        help="Root sum of squares of all uncertainty components"
    )
    
    v_eff_display = "∞" if math.isinf(V_eff) else f"{V_eff:.2f}"
    st.metric(
        "Effective Degrees of Freedom (ν_eff)",
        v_eff_display,
        help="Welch-Satterthwaite effective degrees of freedom. Since V2–V6 are infinite, ν_eff = uc⁴ / (U1⁴ / V1)"
    )
    
    st.metric(
        "Expanded Uncertainty (U)", 
        f"{final_expanded_uncertainty:.6f}",
        delta=f"k = {COVERAGE_FACTOR}" + (" | BMC Floor Applied" if bmc_applied else ""),
        help=f"Combined uncertainty multiplied by coverage factor (k={COVERAGE_FACTOR}). BMC floor = {BMC_FLOOR}%"
    )

# Detailed breakdown
st.markdown("---")
st.subheader("📋 Detailed Calculation Breakdown")

with st.expander("View Calculation Details"):
    st.markdown("### Input Parameters")
    st.write(f"- **Error Readings:** {error_readings}")
    st.write(f"- **Reference Standard Accuracy:** {ref_standard_accuracy}%")
    st.write(f"- **Certificate Uncertainty:** {certificate_uncertainty}%")
    st.write(f"- **DUC Resolution:** {duc_resolution}")
    st.write(f"- **Temperature Difference:** {temp_difference}°C")
    st.write(f"- **Age Factor:** {age_factor}%/year")
    st.write(f"- **Years in Service:** {years_in_service} years")
    st.write(f"- **AC Type:** {ac_type}")
    st.write(f"- **Measurement Mode:** {measurement_mode}")
    st.write(f"- **Voltage:** {voltage} V")
    st.write(f"- **Current:** {current} A")
    st.write(f"- **Power Factor:** {power_factor} ({'Lagging' if pf_type == 'Lag' else ('Leading' if pf_type == 'Lead' else 'Unity')})")
    st.write(f"- **Time Duration:** {time_hours} hours")
    st.write(f"- **Real Power:** {real_power_kw:.3f} kW (calculated using {power_formula})")
    st.write(f"- **Reactive Power:** {reactive_power_kvar:.3f} kVAr (calculated using {reactive_formula})")
    st.write(f"- **Apparent Power:** {apparent_power_kva:.3f} kVA (calculated using {apparent_formula})")
    st.write(f"- **Real Energy:** {real_energy_kwh:.3f} kWh")
    st.write(f"- **Reactive Energy:** {reactive_energy_kvarh:.3f} kVArh")
    st.write(f"- **Apparent Energy:** {apparent_energy_kvah:.3f} kVAh")
    
    st.markdown("### Formulas Used")
    st.latex(r"U_1 = \\sigma_{readings}")
    st.latex(r"U_2 = \\frac{Reference\;Accuracy}{\\sqrt{3}}")
    st.latex(r"U_3 = \\frac{Certificate\;Uncertainty}{2}")
    st.latex(r"U_4 = \\frac{DUC\;Resolution}{2\\sqrt{3}}")
    st.latex(r"U_5 = \\frac{0.000025 \times Temp\;Difference}{\\sqrt{3}}")
    st.latex(r"U_6 = \\frac{Energy\;Drift}{\\sqrt{3}}")
    st.latex(r"u_c = \\sqrt{U_1^2 + U_2^2 + U_3^2 + U_4^2 + U_5^2 + U_6^2}")
    st.latex(r"\nu_{eff} = \frac{u_c^4}{\sum_i \frac{U_i^4}{\nu_i}} = \frac{u_c^4}{\frac{U_1^4}{\nu_1}} \quad (\nu_2\text{–}\nu_6 = \infty)")
    st.latex(r"U = u_c \times k \quad (k=2)")
    
    st.markdown("### Component Contributions")
    contributions = {
        "U1 (Repeatability)": U1**2,
        "U2 (Ref Standard)": U2**2,
        "U3 (Certificate)": U3**2,
        "U4 (Resolution)": U4**2,
        "U5 (Temp Drift)": U5**2,
        "U6 (Energy Drift - Age)": U6**2
    }
    
    total_variance = sum(contributions.values())
    
    for component, variance in contributions.items():
        if total_variance > 0:
            percentage = (variance / total_variance) * 100
            st.write(f"- **{component}:** {variance:.8f} ({percentage:.2f}% contribution)")
        else:
            st.write(f"- **{component}:** {variance:.8f}")

# Excel Export Function
def create_excel_report():
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Uncertainty Calculation"
    
    # Define styles
    header_fill = PatternFill(start_color="0068C9", end_color="0068C9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "Uncertainty Calculation Worksheet - Electrical Lab"
    cell.font = title_font
    cell.alignment = center_alignment
    row += 1
    
    ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 2
    
    # Input Parameters Section
    ws[f'A{row}'] = "INPUT PARAMETERS"
    ws[f'A{row}'].font = bold_font
    row += 1

    ws[f'A{row}'] = "Date of Calibration"
    ws[f'B{row}'] = calibration_date.strftime('%Y-%m-%d') if calibration_date is not None else ""
    row += 1

    ws[f'A{row}'] = "Temperature as per Calibration Certificate of Reference Standard (°C)"
    ws[f'B{row}'] = temp_cal_certificate
    row += 1

    ws[f'A{row}'] = "Room Temperature During Calibration (°C)"
    ws[f'B{row}'] = room_temp_calibration
    row += 1

    ws[f'A{row}'] = "Reference Standard - Make"
    ws[f'B{row}'] = ref_make
    row += 1

    ws[f'A{row}'] = "Reference Standard - Model"
    ws[f'B{row}'] = ref_model
    row += 1

    ws[f'A{row}'] = "Reference Standard - Sl No."
    ws[f'B{row}'] = ref_sl_no
    row += 1

    ws[f'A{row}'] = "Reference Standard - Accuracy Class"
    ws[f'B{row}'] = ref_accuracy_class
    row += 1

    ws[f'A{row}'] = "Reference Standard - Date of Due Calibration"
    ws[f'B{row}'] = ref_due_date.strftime('%Y-%m-%d') if ref_due_date is not None else ""
    row += 1

    ws[f'A{row}'] = "DUC - Make"
    ws[f'B{row}'] = duc_make
    row += 1

    ws[f'A{row}'] = "DUC - Model"
    ws[f'B{row}'] = duc_model
    row += 1

    ws[f'A{row}'] = "DUC - Sl No."
    ws[f'B{row}'] = duc_sl_no
    row += 1

    ws[f'A{row}'] = "DUC - Accuracy Class"
    ws[f'B{row}'] = duc_accuracy_class
    row += 1

    ws[f'A{row}'] = "DUC - Date of Due Calibration"
    ws[f'B{row}'] = duc_due_date.strftime('%Y-%m-%d') if duc_due_date is not None else ""
    row += 1

    ws[f'A{row}'] = "Reference Standard Accuracy (%)"
    ws[f'B{row}'] = ref_standard_accuracy
    row += 1
    
    ws[f'A{row}'] = "Certificate Uncertainty (%)"
    ws[f'B{row}'] = certificate_uncertainty
    row += 1
    
    ws[f'A{row}'] = "DUC Resolution"
    ws[f'B{row}'] = duc_resolution
    row += 1
    
    ws[f'A{row}'] = "Temperature Difference (°C)"
    ws[f'B{row}'] = temp_difference
    row += 1
    
    ws[f'A{row}'] = "Age Factor (%/year)"
    ws[f'B{row}'] = age_factor
    row += 1
    
    ws[f'A{row}'] = "Years in Service"
    ws[f'B{row}'] = years_in_service
    row += 1
    
    ws[f'A{row}'] = "AC Type"
    ws[f'B{row}'] = ac_type
    row += 1
    
    ws[f'A{row}'] = "Measurement Mode"
    ws[f'B{row}'] = measurement_mode
    row += 1
    
    ws[f'A{row}'] = "Voltage (V)"
    ws[f'B{row}'] = voltage
    row += 1
    
    ws[f'A{row}'] = "Current (A)"
    ws[f'B{row}'] = current
    row += 1
    
    ws[f'A{row}'] = "Power Factor"
    ws[f'B{row}'] = power_factor
    row += 1
    
    ws[f'A{row}'] = "Power Factor Type"
    ws[f'B{row}'] = "Lagging" if pf_type == "Lag" else ("Leading" if pf_type == "Lead" else "Unity")
    row += 1
    
    ws[f'A{row}'] = "Time Duration (hours)"
    ws[f'B{row}'] = time_hours
    row += 1
    
    ws[f'A{row}'] = "Real Power (kW)"
    ws[f'B{row}'] = real_power_kw
    row += 1
    
    ws[f'A{row}'] = "Reactive Power (kVAr)"
    ws[f'B{row}'] = reactive_power_kvar
    row += 1
    
    ws[f'A{row}'] = "Apparent Power (kVA)"
    ws[f'B{row}'] = apparent_power_kva
    row += 1
    
    ws[f'A{row}'] = "Real Energy (kWh)"
    ws[f'B{row}'] = real_energy_kwh
    row += 1
    
    ws[f'A{row}'] = "Reactive Energy (kVArh)"
    ws[f'B{row}'] = reactive_energy_kvarh
    row += 1
    
    ws[f'A{row}'] = "Apparent Energy (kVAh)"
    ws[f'B{row}'] = apparent_energy_kvah
    row += 2
    
    # Error Readings Section
    ws[f'A{row}'] = "ERROR READINGS"
    ws[f'A{row}'].font = bold_font
    row += 1
    
    ws[f'A{row}'] = "Reading #"
    ws[f'B{row}'] = "Error Value"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'].font = bold_font
    row += 1
    
    for i, reading in enumerate(error_readings, 1):
        ws[f'A{row}'] = f"Reading {i}"
        ws[f'B{row}'] = reading
        row += 1
    
    row += 1
    
    # Uncertainty Budget Table
    ws[f'A{row}'] = "UNCERTAINTY BUDGET"
    ws[f'A{row}'].font = bold_font
    row += 1
    
    # Headers
    headers = ["Component", "Description", "Type", "Distribution", "Value", "Divisor", "Standard Uncertainty (ui)", "Degrees of Freedom (ν)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    row += 1
    
    # Data rows
    for _, budget_row in uncertainty_budget.iterrows():
        ws[f'A{row}'] = budget_row['Component']
        ws[f'B{row}'] = budget_row['Description']
        ws[f'C{row}'] = budget_row['Type']
        ws[f'D{row}'] = budget_row['Distribution']
        ws[f'E{row}'] = budget_row['Value']
        ws[f'F{row}'] = budget_row['Divisor']
        ws[f'G{row}'] = budget_row['Standard Uncertainty (ui)']
        ws[f'H{row}'] = budget_row['Degrees of Freedom (ν)']
        row += 1
    
    row += 1
    
    # Final Results Section
    ws[f'A{row}'] = "FINAL RESULTS"
    ws[f'A{row}'].font = bold_font
    row += 1
    
    ws[f'A{row}'] = "Average Error (%)"
    ws[f'B{row}'] = average_error
    row += 1
    
    ws[f'A{row}'] = "Combined Uncertainty - uc (%)"
    ws[f'B{row}'] = uc
    row += 1
    
    ws[f'A{row}'] = "Effective Degrees of Freedom (ν_eff)"
    ws[f'B{row}'] = "Infinity" if math.isinf(V_eff) else round(V_eff, 4)
    row += 1
    
    ws[f'A{row}'] = "Expanded Uncertainty - U (k={COVERAGE_FACTOR}) (%)"
    ws[f'B{row}'] = final_expanded_uncertainty
    row += 1
    
    if bmc_applied:
        ws[f'A{row}'] = "BMC Floor Applied"
        ws[f'B{row}'] = f"Calculated: {expanded_uncertainty:.6f}% | Final: {BMC_FLOOR}%"
        ws[f'A{row}'].font = Font(italic=True, color="FF6600")
        row += 1
    
    row += 1
    
    # Final Result Statement
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = f"Result: {average_error:.4f}% ± {final_expanded_uncertainty:.4f}% (k={COVERAGE_FACTOR})"
    cell.font = Font(bold=True, size=14, color="0068C9")
    cell.alignment = center_alignment
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 25
    
    wb.save(output)
    output.seek(0)
    return output

# PDF Export Function
def create_pdf_report():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Helper function to clean special characters
    def clean_text(text):
        text = str(text)
        text = text.replace("±", "+/-")
        text = text.replace("√", "sqrt")
        text = text.replace("°", " deg")
        text = text.replace("₁", "1")
        text = text.replace("₂", "2")
        text = text.replace("₃", "3")
        text = text.replace("₄", "4")
        text = text.replace("₅", "5")
        text = text.replace("₆", "6")
        text = text.replace("φ", "phi")
        text = text.replace("×", "x")
        return text
    
    # Header
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Uncertainty Calculation Worksheet - MTSL Palakkad", ln=True, align="C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 6, "Meter Testing & Standards Laboratory, Palakkad", ln=True, align="C")
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
    pdf.ln(5)
    
    # Section 1 - Input Parameters
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "INPUT PARAMETERS", ln=True)
    pdf.set_font("Arial", "", 10)

    pdf.cell(0, 6, f"Date of Calibration: {calibration_date.strftime('%Y-%m-%d') if calibration_date is not None else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"Temperature as per Calibration Certificate of Reference Standard: {temp_cal_certificate:.2f} deg C", ln=True)
    pdf.cell(0, 6, f"Room Temperature During Calibration: {room_temp_calibration:.2f} deg C", ln=True)
    pdf.ln(2)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Reference Standard Details:", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"  Make: {ref_make if ref_make else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Model: {ref_model if ref_model else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Sl No.: {ref_sl_no if ref_sl_no else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Accuracy Class: {ref_accuracy_class if ref_accuracy_class else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Date of Due Calibration: {ref_due_date.strftime('%Y-%m-%d') if ref_due_date is not None else 'N/A'}", ln=True)
    pdf.ln(2)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Device Under Calibration Details:", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"  Make: {duc_make if duc_make else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Model: {duc_model if duc_model else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Sl No.: {duc_sl_no if duc_sl_no else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Accuracy Class: {duc_accuracy_class if duc_accuracy_class else 'N/A'}", ln=True)
    pdf.cell(0, 6, f"  Date of Due Calibration: {duc_due_date.strftime('%Y-%m-%d') if duc_due_date is not None else 'N/A'}", ln=True)
    pdf.ln(2)

    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"Reference Standard Accuracy: {ref_standard_accuracy:.4f}%", ln=True)
    pdf.cell(0, 6, f"Certificate Uncertainty: {certificate_uncertainty:.4f}%", ln=True)
    pdf.cell(0, 6, f"DUC Resolution: {duc_resolution:.4f}", ln=True)
    pdf.cell(0, 6, f"Temperature Difference: {temp_difference:.2f} deg C", ln=True)
    pdf.cell(0, 6, f"Age Factor: {age_factor:.5f}%/year", ln=True)
    pdf.cell(0, 6, f"Years in Service: {years_in_service:.1f} years", ln=True)
    pdf.cell(0, 6, f"AC Type: {clean_text(ac_type)}", ln=True)
    pdf.cell(0, 6, f"Measurement Mode: {measurement_mode}", ln=True)
    pdf.cell(0, 6, f"Voltage: {voltage:.2f} V", ln=True)
    pdf.cell(0, 6, f"Current: {current:.3f} A", ln=True)
    pdf.cell(0, 6, f"Power Factor: {power_factor:.3f} ({'Lagging' if pf_type == 'Lag' else ('Leading' if pf_type == 'Lead' else 'Unity')})", ln=True)
    pdf.cell(0, 6, f"Time Duration: {time_hours:.2f} hours", ln=True)
    pdf.cell(0, 6, f"Real Power: {real_power_kw:.3f} kW (Formula: {clean_text(power_formula)})", ln=True)
    pdf.cell(0, 6, f"Reactive Power: {reactive_power_kvar:.3f} kVAr (Formula: {clean_text(reactive_formula)})", ln=True)
    pdf.cell(0, 6, f"Apparent Power: {apparent_power_kva:.3f} kVA (Formula: {clean_text(apparent_formula)})", ln=True)
    pdf.cell(0, 6, f"Real Energy: {real_energy_kwh:.3f} kWh", ln=True)
    pdf.cell(0, 6, f"Reactive Energy: {reactive_energy_kvarh:.3f} kVArh", ln=True)
    pdf.cell(0, 6, f"Apparent Energy: {apparent_energy_kvah:.3f} kVAh", ln=True)
    pdf.ln(3)
    
    # Error Readings
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Error Readings (10 measurements):", ln=True)
    pdf.set_font("Arial", "", 9)
    
    # Display readings in 2 rows
    for i in range(0, 10, 5):
        readings_line = "  ".join([f"R{j+1}: {error_readings[j]:.4f}" for j in range(i, min(i+5, 10))])
        pdf.cell(0, 5, readings_line, ln=True)
    pdf.ln(3)
    
    # Section 2 - Uncertainty Budget
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "UNCERTAINTY BUDGET", ln=True)
    pdf.ln(2)
    
    # Table headers
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(0, 104, 201)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(18, 7, "Comp.", border=1, fill=True, align="C")
    pdf.cell(20, 7, "Type", border=1, fill=True, align="C")
    pdf.cell(32, 7, "Distribution", border=1, fill=True, align="C")
    pdf.cell(20, 7, "Value", border=1, fill=True, align="C")
    pdf.cell(18, 7, "Divisor", border=1, fill=True, align="C")
    pdf.cell(25, 7, "Std Unc (ui)", border=1, fill=True, align="C")
    pdf.ln()
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 8)
    
    # Table data
    for idx, row in uncertainty_budget.iterrows():
        comp_name = clean_text(row['Component'])
        # Shorten component name for table
        if "U1" in comp_name:
            comp_short = "U1"
        elif "U2" in comp_name:
            comp_short = "U2"
        elif "U3" in comp_name:
            comp_short = "U3"
        elif "U4" in comp_name:
            comp_short = "U4"
        elif "U5" in comp_name:
            comp_short = "U5"
        elif "U6" in comp_name:
            comp_short = "U6"
        else:
            comp_short = comp_name[:10]
        
        pdf.cell(18, 6, comp_short, border=1, align="C")
        pdf.cell(20, 6, clean_text(row['Type']), border=1, align="C")
        pdf.cell(32, 6, clean_text(row['Distribution']), border=1, align="C")
        pdf.cell(20, 6, f"{row['Value']:.6f}", border=1, align="C")
        pdf.cell(18, 6, clean_text(row['Divisor']), border=1, align="C")
        pdf.cell(25, 6, f"{row['Standard Uncertainty (ui)']:.6f}", border=1, align="C")
        pdf.ln()
    
    pdf.ln(3)
    
    # Component Descriptions
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Component Descriptions:", ln=True)
    pdf.set_font("Arial", "", 8)
    
    for idx, row in uncertainty_budget.iterrows():
        comp_name = clean_text(row['Component'])
        description = clean_text(row['Description'])
        pdf.multi_cell(0, 4, f"{comp_name}: {description}")
        pdf.ln(1)
    
    pdf.ln(2)
    
    # Section 3 - Detailed Step-by-Step Calculations
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "DETAILED STEP-BY-STEP CALCULATIONS", ln=True)
    pdf.ln(2)

    # Helper to render a step block: label, formula line, substitution line, result line
    def step_block(label, formula, substitution, result_line):
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 6, label, ln=True)
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 5, f"  Formula     : {formula}", ln=True)
        pdf.cell(0, 5, f"  Substitution: {substitution}", ln=True)
        pdf.cell(0, 5, f"  Result       : {result_line}", ln=True)
        pdf.ln(2)

    # --- Average Error ---
    sum_of_errors = sum(error_readings)
    readings_list = ", ".join([f"{r:.4f}" for r in error_readings])
    step_block(
        "Average Error",
        "Average = sum(Readings) / n",
        f"= ({readings_list}) / {len(error_readings)}",
        f"= {sum_of_errors:.6f} / {len(error_readings)} = {average_error:.6f} %"
    )

    # --- U1 - Repeatability ---
    std_dev = np.std(error_array, ddof=1)
    step_block(
        "U1 - Repeatability  [Type A, Normal distribution]",
        "U1 = Standard Deviation / sqrt(n)",
        f"= {std_dev:.6f} / sqrt({len(error_array)})",
        f"= {std_dev:.6f} / {math.sqrt(len(error_array)):.6f} = {U1:.6f}"
    )

    # --- U2 - Reference Standard ---
    step_block(
        "U2 - Reference Standard Accuracy  [Type B, Rectangular distribution]",
        "U2 = Reference Accuracy / sqrt(3)",
        f"= {ref_standard_accuracy:.6f} / sqrt(3)",
        f"= {ref_standard_accuracy:.6f} / 1.732051 = {U2:.6f} %"
    )

    # --- U3 - Certificate Uncertainty ---
    step_block(
        "U3 - Certificate Uncertainty  [Type B, Normal distribution, k=2]",
        "U3 = Certificate Uncertainty / 2",
        f"= {certificate_uncertainty:.6f} / 2",
        f"= {U3:.6f} %"
    )

    # --- U4 - Resolution ---
    step_block(
        "U4 - DUC Resolution  [Type B, Rectangular distribution]",
        "U4 = DUC Resolution / (2 * sqrt(3))",
        f"= {duc_resolution:.6f} / (2 x 1.732051)",
        f"= {duc_resolution:.6f} / 3.464102 = {U4:.6f}"
    )

    # --- U5 - Temperature Drift ---
    temp_val = TEMPERATURE_COEFFICIENT * temp_difference
    step_block(
        "U5 - Temperature Drift  [Type B, Rectangular distribution]",
        "U5 = (Temp Coeff x Delta_T) / sqrt(3)",
        f"= (0.000025 x {temp_difference:.2f}) / sqrt(3)  =>  {temp_val:.8f} / 1.732051",
        f"= {U5:.8f}"
    )

    # --- U6 - Energy Drift (Age) ---
    total_drift_val = age_factor * years_in_service
    step_block(
        "U6 - Energy Drift (Age of Reference Standard)  [Type B, Rectangular distribution]",
        "U6 = (Age Factor x Years in Service) / sqrt(3)",
        f"= ({age_factor:.7f} x {years_in_service:.1f}) / sqrt(3)  =>  {total_drift_val:.8f} / 1.732051",
        f"= {U6:.8f} %"
    )

    # --- Combined Uncertainty ---
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Combined Uncertainty (uc)  [Root Sum of Squares]", ln=True)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, "  Formula     : uc = sqrt(U1^2 + U2^2 + U3^2 + U4^2 + U5^2 + U6^2)", ln=True)
    pdf.cell(0, 5, f"  U1^2 = {U1**2:.10f}", ln=True)
    pdf.cell(0, 5, f"  U2^2 = {U2**2:.10f}", ln=True)
    pdf.cell(0, 5, f"  U3^2 = {U3**2:.10f}", ln=True)
    pdf.cell(0, 5, f"  U4^2 = {U4**2:.10f}", ln=True)
    pdf.cell(0, 5, f"  U5^2 = {U5**2:.10f}", ln=True)
    pdf.cell(0, 5, f"  U6^2 = {U6**2:.10f}", ln=True)
    sum_sq = U1**2 + U2**2 + U3**2 + U4**2 + U5**2 + U6**2
    pdf.cell(0, 5, f"  Sum of squares = {sum_sq:.12f}", ln=True)
    pdf.cell(0, 5, f"  Result       : uc = sqrt({sum_sq:.12f}) = {uc:.6f} %", ln=True)
    pdf.ln(2)

    # --- Expanded Uncertainty ---
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Expanded Uncertainty (U)  [Coverage Factor k=2, ~95% confidence]", ln=True)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, "  Formula     : U = k x uc", ln=True)
    pdf.cell(0, 5, f"  Substitution: U = {COVERAGE_FACTOR} x {uc:.6f}", ln=True)
    pdf.cell(0, 5, f"  Result       : U = {expanded_uncertainty:.6f} %", ln=True)
    pdf.ln(2)

    # --- Effective Degrees of Freedom (Welch-Satterthwaite) ---
    v_eff_str = "Infinity" if math.isinf(V_eff) else f"{V_eff:.4f}"
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Effective Degrees of Freedom (nu_eff)  [Welch-Satterthwaite equation]", ln=True)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, "  Formula     : V_eff = uc^4 / sum(Ui^4 / Vi)", ln=True)
    pdf.cell(0, 5, "  Note        : V2-V6 are infinite (Type B, known distributions), so their terms = 0", ln=True)
    pdf.cell(0, 5, "  Simplified  : V_eff = uc^4 / (U1^4 / V1)", ln=True)
    pdf.cell(0, 5, f"  V1 = {V1}  (n-1 = 10-1 = 9, from 10 repeatability readings)", ln=True)
    pdf.cell(0, 5, f"  Substitution: V_eff = {uc:.6f}^4 / ({U1:.6f}^4 / {V1})", ln=True)
    pdf.cell(0, 5, f"               = {uc**4:.10f} / {(U1**4)/V1:.10f}", ln=True)
    pdf.cell(0, 5, f"  Result       : V_eff = {v_eff_str}", ln=True)
    pdf.ln(2)

    if bmc_applied:
        pdf.set_font("Arial", "B", 9)
        pdf.set_text_color(255, 102, 0)
        pdf.cell(0, 5, "BMC Floor Applied:", ln=True)
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 5, f"  Calculated U ({expanded_uncertainty:.6f}%) is below the BMC floor ({BMC_FLOOR:.6f}%).", ln=True)
        pdf.cell(0, 5, f"  U_final = max({expanded_uncertainty:.6f}, {BMC_FLOOR}) = {final_expanded_uncertainty:.4f} %", ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    pdf.ln(3)
    
    # Section 4 - Final Result
    pdf.set_font("Arial", "B", 14)
    pdf.set_text_color(0, 104, 201)
    pdf.cell(0, 10, f"FINAL RESULT: {average_error:.4f}% +/- {final_expanded_uncertainty:.4f}% (k={COVERAGE_FACTOR})", ln=True, align="C", border=1)
    pdf.set_text_color(0, 0, 0)
    
    # Footer
    pdf.ln(5)
    pdf.set_font("Arial", "I", 8)
    pdf.cell(0, 5, "Coverage Factor k=2 (95.45% confidence level)", ln=True, align="C")
    
    # Output to BytesIO
    output = BytesIO()
    pdf_output = pdf.output(dest='S').encode('latin-1')
    output.write(pdf_output)
    output.seek(0)
    return output

# Download Button
st.markdown("---")
col_download1, col_download2 = st.columns(2)

with col_download1:
    excel_file = create_excel_report()
    st.download_button(
        label="📥 Download Calculation as Excel",
        data=excel_file,
        file_name=f"uncertainty_calculation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with col_download2:
    pdf_file = create_pdf_report()
    st.download_button(
        label="📄 Download PDF Report",
        data=pdf_file,
        file_name=f"MTSL_Uncertainty_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mime="application/pdf",
        use_container_width=True
    )

# Final Result Display
st.markdown("---")
st.markdown(
    f"""
    <div style='text-align: center; padding: 20px; background-color: #f0f2f6; border-radius: 10px; margin: 20px 0;'>
        <h2 style='color: #0068c9; margin: 0;'>
            <strong>Result: {average_error:.4f}% ± {final_expanded_uncertainty:.4f}% (k={COVERAGE_FACTOR})</strong>
        </h2>
    </div>
    """, 
    unsafe_allow_html=True
)

# Detailed Step-by-Step Calculation
with st.expander("🔍 Show Detailed Step-by-Step Calculation"):
    st.markdown("### Mathematical Proof and Derivation")
    st.markdown("---")
    
    # Average Error
    st.markdown("#### **Average Error**")
    sum_of_errors = sum(error_readings)
    st.latex(r"\text{Average Error} = \frac{\sum_{i=1}^{10} \text{Error}_i}{10}")
    st.latex(f"= \\frac{{{sum_of_errors:.6f}}}{{10}} = {average_error:.6f} \%")
    st.markdown("")
    
    # U1 - Repeatability
    st.markdown("#### **U₁ - Repeatability (Standard Deviation / √n)**")
    st.latex(r"U_1 = \frac{\sigma}{\sqrt{n}} = \frac{1}{\sqrt{n}}\sqrt{\frac{\sum_{i=1}^{n}(x_i - \bar{x})^2}{n-1}}")
    readings_str = ", ".join([f"{r:.4f}" for r in error_readings])
    st.markdown(f"Where readings are: [{readings_str}], n = {len(error_readings)}")
    st.latex(f"U_1 = {U1:.6f}")
    st.markdown("")
    
    # U2 - Reference Standard
    st.markdown("#### **U₂ - Reference Standard Accuracy (uncertainty caused due to the accuracy factor of reference standard used for calibration)**")
    st.latex(r"U_{2} = \\frac{Reference Accuracy}{\\sqrt{3}}")
    st.latex(f"= \\frac{{{ref_standard_accuracy:.6f}}}{{\\sqrt{{3}}}} = \\frac{{{ref_standard_accuracy:.6f}}}{{1.732051}} = {U2:.6f} \%")
    st.markdown("")
    
    # U3 - Certificate
    st.markdown("#### **U₃ - Certificate Uncertainty (uncertainty caused due to the uncertainty reported by the lab where reference standard was calibrated)**")
    st.latex(r"U_{3} = \\frac{Certificate Uncertainty}{2}")
    st.latex(f"= \\frac{{{certificate_uncertainty:.6f}}}{{2}} = {U3:.6f} \%")
    st.markdown("")
    
    # U4 - Resolution
    st.markdown("#### **U₄ - DUC Resolution**")
    st.latex(r"U_{4} = \\frac{DUC Resolution}{2\\sqrt{3}}")
    st.latex(f"= \\frac{{{duc_resolution:.6f}}}{{2 \\times 1.732051}} = \\frac{{{duc_resolution:.6f}}}{{3.464102}} = {U4:.6f}")
    st.markdown("")
    
    # U5 - Temperature Drift
    st.markdown("#### **U₅ - Temperature Drift (uncertainty caused due to the fact that temperature maintained during the calibration by the lab where reference standard was calibrated is different from the reference temperature)**")
    st.latex(r"U_{5} = \\frac{Temp Coeff \times \Delta T}{\\sqrt{3}}")
    temp_value = TEMPERATURE_COEFFICIENT * temp_difference
    st.latex(f"= \\frac{{0.000025 \times {temp_difference:.2f}}}{{\\sqrt{{3}}}} = \\frac{{{temp_value:.8f}}}{{1.732051}} = {U5:.8f}")
    st.markdown("")
    
    # U6 - Energy Drift (Age-Based)
    st.markdown("#### **U₆ - Energy Drift (uncertainty included to account for age of the reference standard)**")
    st.latex(r"U_{drift} = \\frac{Age Factor \times Years}{\\sqrt{3}}")
    st.latex(f"= \\frac{{{age_factor:.5f} \times {years_in_service:.0f}}}{{1.732051}} = {U6:.8f} \%")
    st.markdown("")
    
    st.markdown("---")
    
    # Combined Uncertainty
    st.markdown("#### **Combined Uncertainty (uᴄ)**")
    st.latex(r"u_c = \\sqrt{U_1^2 + U_2^2 + U_3^2 + U_4^2 + U_5^2 + U_6^2}")
    st.latex(f"= \\sqrt{{{U1:.6f}^2 + {U2:.6f}^2 + {U3:.6f}^2 + {U4:.6f}^2 + {U5:.8f}^2 + {U6:.8f}^2}}")
    sum_of_squares = U1**2 + U2**2 + U3**2 + U4**2 + U5**2 + U6**2
    st.latex(f"= \\sqrt{{{sum_of_squares:.12f}}} = {uc:.6f} \%")
    st.markdown("")
    
    # Effective Degrees of Freedom (Welch-Satterthwaite)
    st.markdown("#### **Effective Degrees of Freedom (ν_eff) – Welch-Satterthwaite Equation**")
    st.latex(r"\nu_{eff} = \frac{u_c^4}{\displaystyle\sum_{i=1}^{6} \frac{U_i^4}{\nu_i}}")
    st.markdown(
        "Since ν₂ through ν₆ are infinite (Type B components from known distributions), "
        "their terms evaluate to zero. The formula reduces to:"
    )
    st.latex(r"\nu_{eff} = \frac{u_c^4}{\dfrac{U_1^4}{\nu_1}}")
    st.latex(
        f"= \\frac{{{uc:.6f}^4}}{{\\dfrac{{{U1:.6f}^4}}{{{V1}}}}}"
        f" = \\frac{{{uc**4:.10f}}}{{{(U1**4)/V1:.10f}}}"
        f" = {v_eff_display}"
    )
    st.markdown("")
    
    # Expanded Uncertainty
    st.markdown("#### **Expanded Uncertainty (U)**")
    st.latex(r"U = k \times u_c")
    st.latex(f"= {COVERAGE_FACTOR} \times {uc:.6f} = {expanded_uncertainty:.6f} \%")
    
    # BMC Floor Check
    if bmc_applied:
        st.markdown("#### **⚠️ Best Measurement Capability (BMC) Floor Applied**")
        st.warning(f"**Note:** Calculated U ({expanded_uncertainty:.4f}%) is below the BMC limit. Final Result set to {BMC_FLOOR:.3f}%.")
        st.latex(f"U_{{final}} = \max(U_{{calculated}}, \text{{BMC Floor}}) = \max({expanded_uncertainty:.6f}, {BMC_FLOOR}) = {final_expanded_uncertainty:.4f} \%")
        st.markdown("")
    
    st.markdown("---")
    st.markdown("### **Final Result**")
    st.latex(f"\text{{Result}} = {average_error:.4f}\% \pm {final_expanded_uncertainty:.4f}\% \quad (k={COVERAGE_FACTOR})")


# Footer
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: gray; font-size: 0.9em;'>
    Uncertainty Calculation Worksheet | Electrical Lab | Coverage Factor k=2
    </div>
    """, 
    unsafe_allow_html=True
)